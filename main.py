from flask import Flask, request, jsonify
import pandas as pd
from datetime import timedelta
from io import BytesIO
import base64
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

app = Flask(__name__)

@app.route('/run-diff', methods=['POST'])
def run_diff():
    # Load CSV content from POST body
    try:
        file_data = request.json['csv_content']
        decoded = base64.b64decode(file_data)
        df = pd.read_csv(BytesIO(decoded), parse_dates=["lastused"])
    except Exception as e:
        return jsonify({"error": f"Invalid CSV data: {str(e)}"}), 400

    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    like_column_pairs = [
        ("email", "email-forced"),
        ("phone", "phone-forced"),
        ("title", "title-forced"),
        ("name", "name-forced"),
        ("mobile", "mobile-forced")
    ]
    keep_columns = ["username", "lastused"] + [col for pair in like_column_pairs for col in pair]
    df = df[keep_columns]

    cutoff = pd.Timestamp.now(tz="UTC") - pd.Timedelta(days=60)
    df = df[df["lastused"].notna() & (df["lastused"] >= cutoff)].drop(columns=["lastused"])

    differences = pd.DataFrame()
    for original, forced in like_column_pairs:
        if original in df.columns and forced in df.columns:
            mask = df[original] != df[forced]
            diff_rows = df[mask].copy()
            diff_rows["field"] = original
            diff_rows["original_value"] = df.loc[mask, original]
            diff_rows["forced_value"] = df.loc[mask, forced]
            differences = pd.concat([differences, diff_rows], ignore_index=True)

    if "username" in differences.columns:
        differences = differences.sort_values("username")

    # Output CSV
    csv_bytes = differences.to_csv(index=False).encode()
    csv_b64 = base64.b64encode(csv_bytes).decode()

    # Output Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Field Differences"

    for r_idx, row in enumerate(dataframe_to_rows(differences, index=False, header=True), 1):
        ws.append(row)
        if r_idx == 1: continue
        orig = ws.cell(row=r_idx, column=ws.max_column - 1)
        forced = ws.cell(row=r_idx, column=ws.max_column)
        orig.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        forced.fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

    table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName="ForcedDifferences", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    table.tableStyleInfo = style
    ws.add_table(table)

    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    excel_io = BytesIO()
    wb.save(excel_io)
    excel_b64 = base64.b64encode(excel_io.getvalue()).decode()

    return jsonify({
        "output_csv_name": f"output2_{timestamp}.csv",
        "output_excel_name": f"output2_{timestamp}.xlsx",
        "csv_base64": csv_b64,
        "excel_base64": excel_b64
    })

if __name__ == '__main__':
    app.run(host="0.0.0.0", port=8080)
